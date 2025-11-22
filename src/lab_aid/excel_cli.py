"""Excel ブックから Lab-Aid の計算を実行するためのコマンドライン補助モジュール。"""

from __future__ import annotations

import argparse
import sys
from collections.abc import Iterable
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from .engine import evaluate

DEFAULT_WORKBOOK = Path("windows") / "lab_aid_input.xlsx"
TOP_HEADERS = ["入力", "入力", "入力", "出力", "出力", "出力"]
COLUMN_HEADERS = [
    "計算式タイプ",
    "計算式",
    "変数",
    "生データ",
    "編集後",
    "報告値",
    "status",
]


def _to_text(value: object | None) -> str:
    """Excel セルへ書き出すテキストに変換する。

    Args:
        value: セルに格納される値。

    Returns:
        `None` の場合は空文字列、それ以外は文字列化した値。
    """
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return f"{value}"
    return str(value)


def _normalize_multiline(value: object | None) -> str:
    """セル内改行を LF に揃える。

    Args:
        value: Excel セルから取得した値。

    Returns:
        改行コードを LF に統一した文字列。
    """
    text = _to_text(value)
    return text.replace("\r\n", "\n").replace("\r", "\n")


def ensure_template(path: Path) -> bool:
    """テンプレートとなる Excel ブックを必要に応じて作成する。

    Args:
        path: テンプレートを配置するパス。

    Returns:
        新規作成した場合は ``True``、既存ファイルがあれば ``False``。
    """

    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "LabAid"

    for col, header in enumerate(TOP_HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)
    ws.cell(row=1, column=7, value="結果")

    # 見出しを見やすくするために 1 行目をセル結合（A1-C1, D1-F1）
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=6)

    for col, header in enumerate(COLUMN_HEADERS, start=1):
        ws.cell(row=2, column=col, value=header)

    # サンプル行
    ws.cell(row=3, column=1, value="E")
    ws.cell(row=3, column=2, value="this = #A * 2\nprint(this, 'ED')")
    ws.cell(row=3, column=3, value="A=5")

    # 列幅とテキスト形式を設定
    for column in range(1, 8):
        col_letter = get_column_letter(column)
        ws.column_dimensions[col_letter].width = 24
        for row in (1, 2, 3):
            cell = ws.cell(row=row, column=column)
            cell.number_format = "@"
    wb.save(path)
    print(f"[情報] 入力テンプレートを作成しました: {path}")
    return True


def _iter_data_rows(ws) -> Iterable[int]:
    """入力が存在するデータ行番号をイテレートする。

    Args:
        ws: 解析対象のワークシート。

    Yields:
        入力列のいずれかに値が存在する行番号。
    """
    max_row = ws.max_row
    for row in range(3, max_row + 1):
        if any(
            _to_text(ws.cell(row=row, column=col).value).strip() for col in (1, 2, 3)
        ):
            yield row


def _evaluate_row(ws, row: int) -> tuple[str | None, str | None, str | None, str]:
    """1 行分の入力を評価し、結果を返す。

    Args:
        ws: 評価対象のワークシート。
        row: 評価する行番号。

    Returns:
        `(raw, edited, reported, status)` のタプル。
    """
    calc_type = _to_text(ws.cell(row=row, column=1).value).strip()
    script = _normalize_multiline(ws.cell(row=row, column=2).value)
    inputs = _normalize_multiline(ws.cell(row=row, column=3).value)

    if not calc_type:
        return None, None, None, "行が空です (calc_type が未入力)"

    try:
        raw, edited, reported = evaluate(calc_type, script, inputs)
        status = "OK"
    except Exception as exc:  # pragma: no cover - unexpected path
        raw = edited = reported = None
        status = f"ERROR: {exc}"

    return raw, edited, reported, status


def _write_text_cell(ws, row: int, column: int, value: object | None) -> None:
    """テキスト形式のセルを設定する。

    Args:
        ws: 書き込み対象のワークシート。
        row: 対象行番号。
        column: 対象列番号。
        value: 設定する値。
    """
    cell = ws.cell(row=row, column=column, value=_to_text(value))
    cell.number_format = "@"


def _record_results(ws, row: int, raw, edited, reported, status: str) -> None:
    """評価結果をワークシートへ記録する。

    Args:
        ws: 書き込み対象のワークシート。
        row: 対象行番号。
        raw: 生データ列へ記録する値。
        edited: 編集後列へ記録する値。
        reported: 報告値列へ記録する値。
        status: ステータス列へ記録する文字列。
    """
    _write_text_cell(ws, row, 4, raw)
    _write_text_cell(ws, row, 5, edited)
    _write_text_cell(ws, row, 6, reported)
    _write_text_cell(ws, row, 7, status)


def main(argv: list[str] | None = None) -> int:
    """Excel ベースの Lab-Aid 計算を一括実行するエントリーポイント。

    Args:
        argv: コマンドライン引数リスト。``None`` の場合は `sys.argv` を使用。

    Returns:
        成功時は 0、評価対象がない場合は 1 を返す。
    """
    parser = argparse.ArgumentParser(
        description=(
            "Excel ブックに記載された Lab-Aid の計算条件を読み込み、"
            "評価結果を書き戻します。"
        )
    )
    parser.add_argument(
        "workbook",
        nargs="?",
        default=str(DEFAULT_WORKBOOK),
        help="Excel ファイルのパス",
    )
    parser.add_argument(
        "--create-template",
        action="store_true",
        help="テンプレートだけ作成して終了",
    )
    args = parser.parse_args(argv)

    workbook_path = Path(args.workbook).resolve()

    template_created = ensure_template(workbook_path)
    if args.create_template:
        return 0

    wb = load_workbook(workbook_path)
    ws = wb.active

    data_rows = list(_iter_data_rows(ws))
    if not data_rows:
        print(
            "[警告] 評価対象の行が見つかりませんでした。テンプレートを編集して再実行してください。"
        )
        return 1

    print(f"[情報] {len(data_rows)} 行の評価を開始します。")
    for row in data_rows:
        raw, edited, reported, status = _evaluate_row(ws, row)
        _record_results(ws, row, raw, edited, reported, status)
        print(f"  行 {row}: {status}")

    wb.save(workbook_path)
    print(f"[情報] 結果を保存しました: {workbook_path}")
    if template_created:
        print("[ヒント] テンプレートにデータを入力して再実行してください。")
    return 0


if __name__ == "__main__":  # pragma: no cover - CLI entry point
    sys.exit(main())
